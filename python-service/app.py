import sys
import os
import json
import math
from io import BytesIO
from openpyxl import load_workbook
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

# --- Configuration ---
app = Flask(__name__)
# Enable CORS for all routes for frontend access
CORS(app) 
# Simple in-memory storage for raw uploaded data
DATA_STORE = {} 
DATA_ID_COUNTER = 0

# ==========================================
# Section A: Helper Utilities (from utils.py)
# ==========================================

def clean_header(header):
    """Normalizes header string."""
    return str(header).strip().lower() if header else ""

def find_col_index(headers, candidates):
    """Finds the 0-based index of a column header."""
    normalized_headers = [clean_header(h) for h in headers]
    for candidate in candidates:
        if candidate in normalized_headers:
            return normalized_headers.index(candidate)
    return -1

def split_roll_branch(s):
    """Splits student string into roll and branch."""
    if not s: return {"roll": "", "branch": ""}
    s = str(s).strip()
    parts = s.split()
    roll = parts[0]
    branch = " ".join(parts[1:]) if len(parts) > 1 else ""
    return {"roll": roll, "branch": branch}

# ==========================================
# Section B: Core Seating Logic (from logic.py)
# ==========================================

def generate_seating_plan(students, rooms, pattern='standard'):
    """Allocates students to rooms based on the specific pattern."""
    
    student_queue = []
    for pair in students:
        if pair.get('s1'):
            student_queue.append({'val': pair['s1'], 'orig': 'Series 1', 'id': pair['id']})
        if pair.get('s2'):
            student_queue.append({'val': pair['s2'], 'orig': 'Series 2', 'id': pair['id']})

    processed_rooms = [r.copy() for r in rooms]
    queue_idx = 0
    total_students = len(student_queue)

    for room in processed_rooms:
        rows, cols = room['rows'], room['cols']
        coords = [] # (row_index, col_index, capacity, [optional_side])

        # --- Pattern Logic Switch ---
        if pattern == 'columnar':
            # Fill column by column from top to bottom.
            # Implemented as Seat-Filling: All Lefts in Col 0, then All Rights in Col 0, etc.
            # This ensures Strict Vertical filling.
            for c in range(cols):
                # Left Side
                for r in range(rows): coords.append((r, c, 1, 'left'))
                # Right Side
                for r in range(rows): coords.append((r, c, 1, 'right'))
                
        elif pattern == 'snake-vertical':
            # Alternate columns should be reversed.
            for c in range(cols):
                # Standard Vertical: Down
                col_seats = []
                for r in range(rows):
                    col_seats.append((r, c, 1, 'left'))
                    col_seats.append((r, c, 1, 'right'))
                
                if c % 2 == 1:
                    col_seats.reverse()
                
                coords.extend(col_seats)

        elif pattern == 'checkerboard':
            # Students placed on alternating seats (Anti-Cheat).
            # Seat index in row: 2*c (Left), 2*c+1 (Right)
            for r in range(rows):
                for c in range(cols):
                    # Check absoluate seat positions
                    # Left Seat
                    if (r + 0 + c * 2) % 2 == 0:
                        coords.append((r, c, 1, 'left'))
                    # Right Seat
                    if (r + 1 + c * 2) % 2 == 0:
                        coords.append((r, c, 1, 'right'))

        elif pattern == 'single':
            # Each seat (desk) should hold only one student.
            for r in range(rows):
                for c in range(cols): 
                    coords.append((r, c, 1, 'left'))

        elif pattern == 'alternate-rows':
            # Even rows different capacity than odd.
            # Row 0=2, Row 1=1
            for r in range(rows):
                capacity = 2 if r % 2 == 0 else 1
                for c in range(cols): coords.append((r, c, capacity))

        elif pattern == 'hybrid':
            # Columns alternate capacity.
            # Col 0=2, Col 1=1
            for c in range(cols):
                capacity = 2 if c % 2 == 0 else 1
                for r in range(rows): coords.append((r, c, capacity))

        elif pattern == 'staggered':
            # Rows alternate seating on left and right.
            for r in range(rows):
                side = 'left' if r % 2 == 0 else 'right'
                for c in range(cols): coords.append((r, c, 1, side))

        else: # Standard (Z) & Snake (S)
            for r in range(rows):
                row_coords = [(r, c, 2) for c in range(cols)]
                if pattern == 'snake' and r % 2 == 1: row_coords.reverse()
                coords.extend(row_coords)

        # --- Fill Grid ---
        grid = [[{'left': None, 'right': None} for _ in range(cols)] for _ in range(rows)]
        assignments_count = 0

        for item in coords:
            if len(item) == 4:
                r, c, cap, side = item
            else:
                r, c, cap = item
                side = 'left'

            if queue_idx < total_students:
                if side == 'left':
                    s1_obj = student_queue[queue_idx]
                    grid[r][c]['left'] = s1_obj
                    queue_idx += 1
                    assignments_count += 1
                    
                    if cap == 2 and queue_idx < total_students:
                        s2_obj = student_queue[queue_idx]
                        grid[r][c]['right'] = s2_obj
                        queue_idx += 1
                        assignments_count += 1
                elif side == 'right':
                    s1_obj = student_queue[queue_idx]
                    grid[r][c]['right'] = s1_obj
                    queue_idx += 1
                    assignments_count += 1
        
        room['grid'] = grid
        room['assigned_count'] = assignments_count

    unallocated = student_queue[queue_idx:]
    
    return processed_rooms, unallocated

def convert_grid_to_assigned_data(processed_rooms): # Helper to prevent EOF issues if cut off
    """Converts grid-based room data to the assignedData format expected by PDF/Frontend."""
    assigned_data = {}
    for room in processed_rooms:
        pairs = []
        summary = {}
        
        # Iterate grid to extract pairs
        for r in range(room['rows']):
            for c in range(room['cols']):
                desk = room['grid'][r][c]
                if desk:
                    s1_obj = desk.get('left')
                    s2_obj = desk.get('right')
                    s1 = s1_obj.get('val', '') if s1_obj else ''
                    s2 = s2_obj.get('val', '') if s2_obj else ''
                    
                    if s1 or s2:
                        pairs.append({'s1': s1, 's2': s2})
                        
                        # Update summary
                        if s1:
                            branch = split_roll_branch(s1)['branch']
                            summary[branch] = summary.get(branch, 0) + 1
                        if s2:
                            branch = split_roll_branch(s2)['branch']
                            summary[branch] = summary.get(branch, 0) + 1
                            
        assigned_data[room['name']] = {
            'pairs': pairs,
            'summary': summary
        }
    return assigned_data

# ==========================================
# Section C: API Endpoints
# ==========================================

class ContextSetter(Flowable):
    """
    A invisible Flowable that updates the canvas context (room info) 
    whenever it is processed during the document build.
    This allows page headers to change dynamically between rooms.
    """
    def __init__(self, context):
        Flowable.__init__(self)
        self.context = context

    def draw(self):
        # Update the canvas with the current context
        self.canv.header_context = self.context
    
    def wrap(self, availWidth, availHeight):
        return (0, 0) # Takes up no space

def draw_header(canvas, doc):
    """
    Draws the fixed header on every page using the current context 
    stored in the canvas.
    """
    canvas.saveState()
    
    # Retrieve context or use defaults
    ctx = getattr(canvas, 'header_context', {})
    college = ctx.get('college', "GALGOTIAS EDUCATIONAL INSTITUTIONS, GREATER NOIDA")
    exam = ctx.get('exam', "1st CAE (ODD-2025-26)")
    room_name = ctx.get('room_name', "")
    student_count = ctx.get('student_count', 0)
    
    width, height = doc.pagesize
    
    # 1. College Name (Centered, Bold, Largest)
    canvas.setFont("Helvetica-Bold", 16)
    canvas.drawCentredString(width / 2.0, height - 50, college.upper())
    
    # 2. Exam Name (Centered, Slightly Smaller)
    canvas.setFont("Helvetica", 12)
    canvas.drawCentredString(width / 2.0, height - 70, exam)
    
    # 3. Seating Plan Line (Centered)
    canvas.setFont("Helvetica-Bold", 12)
    canvas.drawCentredString(width / 2.0, height - 90, "SEATING PLAN")
    
    # 4. Room & Student Info Line (Left & Right aligned on same line)
    y_pos = height - 120
    canvas.setFont("Helvetica-Bold", 11)
    
    # Room Number (Left aligned with margin)
    left_margin = doc.leftMargin
    canvas.drawString(left_margin, y_pos, f"Room Number: {room_name}")
    
    # Total Students (Right aligned with margin)
    # We calculate string width to align it properly to the right margin
    count_text = f"Total Students: {student_count}"
    text_width = canvas.stringWidth(count_text, "Helvetica-Bold", 11)
    right_margin = width - doc.rightMargin
    canvas.drawString(right_margin - text_width, y_pos, count_text)
    
    # 5. White Board Indicator (Centered, Separated)
    canvas.setFont("Helvetica", 9)
    canvas.drawCentredString(width / 2.0, y_pos - 25, "↑↑↑↑↑↑↑↑↑ White Board ↑↑↑↑↑↑↑↑↑")
    
    canvas.restoreState()

def generate_pdf_internal(rooms, assigned_data, unallocated=None):
    buffer = BytesIO()
    # Increased topMargin to 2.2 inch to accommodate the header
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=2.2*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], alignment=1, fontSize=14, spaceAfter=6)
    
    # Custom Style for Table Cells (Crucial for fitting content)
    # alignment=1 is CENTER
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontSize=7, leading=8, alignment=1) 

    for room in rooms:
        # Get pairs for this room
        room_data = assigned_data.get(room['name'], {})
        pairs = room_data.get('pairs', [])
        
        # Calculate Logic
        student_count = 0
        for p in pairs:
            if p.get('s1'): student_count += 1
            if p.get('s2'): student_count += 1
            
        # --- CRITICAL: Inject Context for Header ---
        # Prior to this room's content, we set the canvas context
        context = {
            'college': room.get('college', "GALGOTIAS EDUCATIONAL INSTITUTIONS, GREATER NOIDA"),
            'exam': room.get('exam', "1st CAE (ODD-2025-26)"),
            'room_name': room['name'],
            'student_count': student_count
        }
        elements.append(ContextSetter(context))
        
        # No inline headers needed anymore!
        
        if not pairs:
             elements.append(Paragraph("(No students assigned to this room)", styles['Normal']))
             elements.append(PageBreak())
             continue

        # --- Seating Grid (Multi-Column) ---
        num_super_cols = 4
        rows_per_col = math.ceil(len(pairs) / num_super_cols)
        
        # Grid Setup
        table_grid = [['' for _ in range(12)] for _ in range(rows_per_col + 1)]
        
        headers = ['Seat No.', 'Roll Series 1', 'Roll Series 2']
        for i in range(num_super_cols):
            base_col = i * 3
            table_grid[0][base_col] = headers[0]
            table_grid[0][base_col+1] = headers[1]
            table_grid[0][base_col+2] = headers[2]

        for i, pair in enumerate(pairs):
            col_idx = i // rows_per_col
            row_idx = i % rows_per_col
            
            if col_idx >= num_super_cols:
                break 
            
            base_col = col_idx * 3
            
            # Seat Number
            table_grid[row_idx + 1][base_col] = Paragraph(str(i + 1), cell_style)
            
            s1_raw = pair.get('s1', '')
            s2_raw = pair.get('s2', '')
            s1_parts = split_roll_branch(s1_raw)
            s2_parts = split_roll_branch(s2_raw)
            
            s1_text = f"{s1_parts['roll']}<br/>{s1_parts['branch']}" if s1_raw else ''
            s2_text = f"{s2_parts['roll']}<br/>{s2_parts['branch']}" if s2_raw else ''
            
            table_grid[row_idx + 1][base_col + 1] = Paragraph(s1_text, cell_style)
            table_grid[row_idx + 1][base_col + 2] = Paragraph(s2_text, cell_style)

        col_widths = [0.5*inch, 1.05*inch, 1.05*inch] * 4

        t = Table(table_grid, colWidths=col_widths, repeatRows=1)

        ts = [
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7), 
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eeeeee')), 
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

        t.setStyle(TableStyle(ts))
        elements.append(t)
        elements.append(Spacer(1, 0.2 * inch))

        # --- Footer Summary ---
        summary = room_data.get('summary', {})
        if summary:
            summary_text = "Branch & Students: " + ", ".join([f"{k}={v}" for k, v in summary.items()])
            
            footer_data = [[Paragraph(summary_text, styles['Normal'])]]
            t_footer = Table(footer_data, colWidths=[10.6*inch])
            t_footer.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(t_footer)

        elements.append(PageBreak())
        
    if unallocated:
        elements.append(Paragraph("Students Pending Allocation (Not Seated)", title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        unallocated_data = [['Roll No', 'Branch', 'Series']]
        for student in unallocated:
            if 'val' in student:
                # Raw format from calculation logic
                details = split_roll_branch(student['val'])
                unallocated_data.append([details['roll'], details['branch'], student.get('orig', '')])
            else:
                # Processed format from Frontend
                unallocated_data.append([student.get('roll', ''), student.get('branch', ''), student.get('orig', '')])
            
        ut = Table(unallocated_data)
        ut.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(ut)

    # Build PDF with Callbacks
    doc.build(elements, onFirstPage=draw_header, onLaterPages=draw_header)
    
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='seating_plan.pdf', mimetype='application/pdf')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handles Excel file upload, parses data, and stores it."""
    global DATA_ID_COUNTER

    if 'file' not in request.files:
        return jsonify({"error": "No file part"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400
    
    if file and file.filename.endswith(('.xlsx', '.xls')):
        try:
            # Read file into memory buffer
            file_bytes = file.read()
            wb = load_workbook(BytesIO(file_bytes), data_only=True)
            sheet = wb.active

            headers = [cell.value for cell in sheet[1]]
            
            # Identify Columns
            idx_roll1 = find_col_index(headers, ["roll no. series-1", "series-1", "roll1"])
            idx_roll2 = find_col_index(headers, ["roll no. series-2", "series-2", "roll2"])
            idx_room = find_col_index(headers, ["room no.", "room"])
            idx_rows = find_col_index(headers, ["rows", "row", "no. of rows"])
            idx_cols = find_col_index(headers, ["columns", "cols", "column"])
            idx_college = find_col_index(headers, ["college name", "college"])
            idx_exam = find_col_index(headers, ["exam name", "exam"])

            if idx_roll1 == -1 or idx_roll2 == -1 or idx_room == -1:
                return jsonify({"error": "Missing required columns: Series-1, Series-2, or Room No."}), 400
            
            students = []
            rooms_map = {}
            total_students = 0
            
            # Fill-down variables
            last_college = ""
            last_exam = ""

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                
                # Student pairs
                s1 = str(row[idx_roll1]).strip() if idx_roll1 < len(row) and row[idx_roll1] else ""
                s2 = str(row[idx_roll2]).strip() if idx_roll2 < len(row) and row[idx_roll2] else ""
                
                if s1 or s2:
                    students.append({'s1': s1, 's2': s2, 'id': i + 2})
                    if s1: total_students += 1
                    if s2: total_students += 1

                # Room configurations (must be unique by name)
                r_name = str(row[idx_room]).strip() if idx_room < len(row) and row[idx_room] else ""
                
                # Extract Metadata with Fill-Down Logic
                current_college = str(row[idx_college]).strip() if idx_college != -1 and idx_college < len(row) and row[idx_college] else ""
                current_exam = str(row[idx_exam]).strip() if idx_exam != -1 and idx_exam < len(row) and row[idx_exam] else ""
                
                if current_college:
                    last_college = current_college
                if current_exam:
                    last_exam = current_exam

                if r_name and r_name not in rooms_map:
                    try:
                        r_rows = int(row[idx_rows]) if idx_rows != -1 and idx_rows < len(row) and row[idx_rows] else 0
                        r_cols = int(row[idx_cols]) if idx_cols != -1 and idx_cols < len(row) and row[idx_cols] else 0
                    except (ValueError, TypeError):
                        r_rows, r_cols = 0, 0

                    if r_rows > 0 and r_cols > 0:
                        rooms_map[r_name] = {
                            'name': r_name, 'rows': r_rows, 'cols': r_cols,
                            'college': last_college,
                            'exam': last_exam
                        }
            
            if not students:
                return jsonify({"error": "No student records found."}), 400
            if not rooms_map:
                return jsonify({"error": "No valid room configurations found (check Row/Col counts)."}), 400

            DATA_ID_COUNTER += 1
            data_id = str(DATA_ID_COUNTER)
            DATA_STORE[data_id] = {
                'students': students, 
                'rooms': list(rooms_map.values()),
                'total_students': total_students
            }
            
            return jsonify({
                "message": "File processed successfully",
                "data_id": data_id,
                "total_students": total_students,
                "room_count": len(rooms_map)
            }), 200

        except Exception as e:
            print(f"Server error during processing: {e}")
            return jsonify({"error": f"An error occurred during file processing: {str(e)}"}), 500

    return jsonify({"error": "Invalid file type. Please upload XLSX or XLS."}), 400

@app.route('/calculate', methods=['POST'])
def calculate_seating():
    """Handles the request to run the seating algorithm based on pattern."""
    data = request.get_json()
    data_id = data.get('data_id')
    pattern = data.get('pattern', 'standard')

    if data_id not in DATA_STORE:
        return jsonify({"error": "Data ID not found. Please upload the file again."}), 404

    raw_data = DATA_STORE[data_id]
    
    try:
        processed_rooms, unallocated = generate_seating_plan(
            raw_data['students'], 
            raw_data['rooms'], 
            pattern
        )
        
        # Prepare unallocated list for frontend (splitting roll/branch for display)
        unallocated_for_display = []
        for student in unallocated:
            details = split_roll_branch(student['val'])
            unallocated_for_display.append({
                'roll': details['roll'],
                'branch': details['branch'],
                'orig': student['orig']
            })

        # Prepare room grids for display (splitting roll/branch for DeskCard)
        for room in processed_rooms:
            for r in range(room['rows']):
                for c in range(room['cols']):
                    desk = room['grid'][r][c]
                    if desk and desk.get('left'):
                        left_obj = desk['left']
                        left_details = split_roll_branch(left_obj.get('val', ''))
                        left_details['orig'] = left_obj.get('orig', '')
                        desk['left'] = left_details
                    if desk and desk.get('right'):
                        right_obj = desk['right']
                        right_details = split_roll_branch(right_obj.get('val', ''))
                        right_details['orig'] = right_obj.get('orig', '')
                        desk['right'] = right_details

        return jsonify({
            "rooms": processed_rooms,
            "unallocated": unallocated_for_display,
            "total_students": raw_data['total_students']
        }), 200

    except Exception as e:
        print(f"Calculation error: {e}")
        return jsonify({"error": "Calculation failed. Check room and student data consistency."}), 500

@app.route('/generate-pdf-from-raw', methods=['POST'])
def generate_pdf_from_raw():
    """Receives raw students and rooms, runs allocation, and generates PDF."""
    try:
        data = request.get_json()
        students_raw = data.get('students', [])
        rooms_raw = data.get('rooms', [])
        
        students = []
        for i, p in enumerate(students_raw):
            students.append({'s1': p.get('s1'), 's2': p.get('s2'), 'id': i})
            
        processed_rooms, unallocated = generate_seating_plan(students, rooms_raw)
        assigned_data = convert_grid_to_assigned_data(processed_rooms)
        
        return generate_pdf_internal(rooms_raw, assigned_data, unallocated)

    except Exception as e:
        print(f"PDF Raw Error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/generate-pdf-from-logic', methods=['POST'])
def generate_pdf_from_logic():
    """Generates a PDF based on the provided seating logic data."""
    try:
        data = request.get_json()
        rooms = data.get('rooms', [])
        assigned_data = data.get('assignedData', {})
        # Note: This endpoint might not have unallocated data if not passed.
        # If the frontend updates to pass it, we can use it.
        unallocated = data.get('unallocated', [])
        # unallocated here might be a list of objects, but generate_pdf_internal expects 
        # objects with 'val' and 'orig' keys if we reuse the logic.
        # However, the frontend likely sends formatted unallocated data.
        # Let's assume for now this endpoint is legacy or will be updated later if needed.
        # But to be safe, let's just pass it if it matches the structure or handle it.
        # Given the complexity, let's focus on /generate-pdf-from-raw which is the primary new flow.
        
        return generate_pdf_internal(rooms, assigned_data, unallocated)

    except Exception as e:
        print(f"PDF Generation Error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/preview-allocation', methods=['POST'])
def preview_allocation():
    """Runs allocation on supplied students/rooms and returns preview data."""
    try:
        payload = request.get_json(silent=True) or {}
        pattern = payload.get('pattern', 'standard')
        students = payload.get('students', [])
        rooms = payload.get('rooms', [])

        # Add validation
        if not students or not rooms:
            return jsonify({"error": "Missing students or rooms data"}), 400

        # Add 'id' field if missing
        for i, student in enumerate(students):
            if 'id' not in student:
                student['id'] = i

        processed_rooms, unallocated = generate_seating_plan(students, rooms, pattern)
        assigned_data = convert_grid_to_assigned_data(processed_rooms)

        unallocated_for_display = []
        for student in unallocated:
            # Fix: student has 'val' key from generate_seating_plan
            val = student.get('val', '')
            details = split_roll_branch(val)
            unallocated_for_display.append({
                'roll': details['roll'],
                'branch': details['branch'],
                'orig': student.get('orig', '')
            })

        # Calculate total students
        total_students = 0
        for student in students:
            if student.get('s1'): total_students += 1
            if student.get('s2'): total_students += 1

        return jsonify({
            "rooms": processed_rooms,
            "assignedData": assigned_data,
            "unallocated": unallocated_for_display,
            "total_students": total_students
        }), 200
        
    except KeyError as e:
        print(f"Preview allocation KeyError: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Missing required field: {str(e)}"}), 400
    except Exception as e:
        print(f"Preview allocation error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Allocation failed: {str(e)}"}), 500
@app.route('/health', methods=['GET'])
def health():
    return jsonify({"status": "ok"}), 200

if __name__ == '__main__':
    # Flask runs on port 5000 by default
    print("Starting Flask server on http://127.0.0.1:5000")
    # Set host='0.0.0.0' to be accessible from outside the container/localhost
    port = int(os.environ.get('PORT', 5001))
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(host='0.0.0.0', port=port, debug=debug_mode)